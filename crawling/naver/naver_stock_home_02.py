# # Library
import requests
from bs4 import BeautifulSoup as bs

import os, re
from datetime import datetime
import pandas as pd

from openpyxl import load_workbook

# 페이지 접속
url = 'https://finance.naver.com/' # 네이버 증권 > 홈

# requests 테스트
# - 페이지 접속 가능 여부확인
#     - 가능할 경우 출력 : <Response [200]>
req = requests.get(url)
print(req)

# 정보 수집
# html = req.content.decode('utf-8') # 한글 깨짐 해결 코드
html = req.text
soup = bs(html, 'html.parser')

# 현재 시간 가져오기
current_time = datetime.now()
# current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
print(current_time)
print(soup.title.text)

# TOP 종목
section_sise_top = soup.select_one('div#content div.section_sise_top')
group_types = section_sise_top.select('div.group_type')

## 데이터 출력

# 전일 대비 정보 추출 함수
def extract_number(text):
    # 숫자 추출
    number_str = re.sub('[^0-9]', '', text)
    if number_str == '':
        return 0  # 빈 문자열인 경우 0 반환
    
    number = int(number_str)
    
    # '하락'이 포함된 경우 음수로 변환
    if '하락' in text:
        number = -number
    
    return number

# 현재 날짜
current_date = current_time.strftime('%Y%m%d')

# 현재 경로 확인
code_path = os.getcwd().replace('\\', '/')
# 수집한 파일 저장할 폴더 생성
crawled_folder_path = os.path.join(code_path, 'crawled_data', 'naver_stock', 'home', current_date)
os.makedirs(crawled_folder_path, exist_ok=True)
# 저장할 파일 경로
file_path = os.path.join(crawled_folder_path, f'naver_stock_home_{current_time.strftime("%Y%m%d_%H%M%S")}.xlsx')

# 탭 정보 : 거래상위, 상승, 하락, 시가총액상위
sheet_name_list = ['거래상위', '상승', '하락', '시가총액상위']
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    for group, sheet_name in zip(group_types, sheet_name_list):
        # 종목 정보 추출
        stock_lines = group.select('tbody tr')
        
        stock_name_list = []
        current_price_list = []
        compare_yesterday_list = []
        updown_ratio_list = []
        for one_stock in stock_lines:
            # 종목명
            stock_name = one_stock.select_one('th a').text

            # 현재가
            current_price = int(one_stock.select_one('td').text.replace(',', ''))

            # 전일대비
            raw_compare_yesterday = one_stock.select('td')[1].text
            compare_yesterday = extract_number(raw_compare_yesterday)

            # 등락률
            raw_updown_ratio = one_stock.select('td')[2].text.strip()
            updown_ratio = float(re.sub('[^0-9.-]', '', raw_updown_ratio))

            # 리스트에 추가
            stock_name_list.append(stock_name)
            current_price_list.append(current_price)
            compare_yesterday_list.append(compare_yesterday)
            updown_ratio_list.append(updown_ratio)
        
        # 데이터프레임 생성
        df = pd.DataFrame(
            {
                '수집시간': current_time,
                '종목명': stock_name_list,
                '현재가': current_price_list,
                '전일대비': compare_yesterday_list,
                '등락률(%)': updown_ratio_list
            }
        )
        df.to_excel(writer, sheet_name=sheet_name, index=False)

# 엑셀 파일 로드
wb = load_workbook(file_path)

# 각 시트에 대해 셀 서식 설정
for sheet_name in sheet_name_list:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):  # '등락률' 열에 대해 서식 설정
        for cell in row:
            cell.number_format = '0.00'

# 엑셀 파일 저장
wb.save(file_path)

print('데이터 저장 경로 :', crawled_folder_path)
print('저장완료')

# # END


